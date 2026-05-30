"""
excel_handler.py — Excel import, export, and sync logic
"""
import os
import io
from datetime import datetime, date
import database

UPLOAD_DIR = 'uploads'

# ─── COLUMN DETECTION ──────────────────────────────────────────────────────────

TICKET_NAMES   = ['service ticket', 'ticket', 'ticket id', 'ticket no', 'cx ticket',
                   'sr no', 'sr number', 'sr', 'request no', 'incident']
QUERY_NAMES    = ['query executed', 'query', 'command', 'sql', 'script', 'query/command',
                   'backend query', 'db query', 'sql query', 'correction']
EXEC_NAMES     = ['executed', 'executed by', 'done by', 'agent', 'user', 'operator',
                   'performed by', 'by', 'exec by', 'staff']
DATE_NAMES     = ['date', 'execution date', 'date executed', 'date done', 'exec date',
                   'correction date', 'date of correction']
STATUS_NAMES   = ['status', 'result', 'outcome']
NOTES_NAMES    = ['notes', 'remarks', 'comments', 'note', 'remark', 'observation']

def _find_col(headers, names):
    """Find column index (0-based) from header dict."""
    for name in names:
        if name in headers:
            return headers[name] - 1
    return None

def _detect_columns(ws):
    """Detect column positions from the first header row."""
    headers = {}
    first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not first_row:
        raise Exception('Excel file appears to be empty.')
    for i, cell in enumerate(first_row, 1):
        if cell is not None:
            headers[str(cell).strip().lower()] = i
    return {
        'ticket':      _find_col(headers, TICKET_NAMES),
        'query':       _find_col(headers, QUERY_NAMES),
        'executed_by': _find_col(headers, EXEC_NAMES),
        'date':        _find_col(headers, DATE_NAMES),
        'status':      _find_col(headers, STATUS_NAMES),
        'notes':       _find_col(headers, NOTES_NAMES),
    }, list(headers.keys())

def _get_cell(row, idx):
    if idx is None or idx >= len(row):
        return None
    val = row[idx]
    if val is None:
        return None
    s = str(val).strip()
    return s if s and s.lower() != 'none' else None

def _parse_date(val):
    if val is None:
        return date.today().isoformat()
    if isinstance(val, (datetime,)):
        return val.strftime('%Y-%m-%d')
    if isinstance(val, date):
        return val.strftime('%Y-%m-%d')
    # Try parsing string
    s = str(val).strip()
    for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y', '%m/%d/%Y', '%d %b %Y', '%d-%b-%Y', '%Y/%m/%d'):
        try:
            return datetime.strptime(s, fmt).strftime('%Y-%m-%d')
        except ValueError:
            continue
    return date.today().isoformat()

def _load_workbook(file):
    try:
        import openpyxl
    except ImportError:
        raise Exception('openpyxl is not installed. Run: pip install openpyxl')
    content = file.read() if hasattr(file, 'read') else file
    return openpyxl.load_workbook(io.BytesIO(content), data_only=True)

def _extract_records(ws, col_map):
    """Extract and validate rows from the worksheet."""
    records = []
    skipped = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(c for c in row if c is not None):
            continue
        ticket = _get_cell(row, col_map['ticket'])
        query  = _get_cell(row, col_map['query'])
        if not ticket or not query:
            skipped += 1
            continue
        records.append({
            'ticket':      ticket,
            'query':       query,
            'executed_by': _get_cell(row, col_map['executed_by']) or 'Imported',
            'date':        _parse_date(_get_cell(row, col_map['date'])),
            'status':      _get_cell(row, col_map['status']) or 'Completed',
            'notes':       _get_cell(row, col_map['notes']) or '',
        })
    return records, skipped

# ─── FILE HEADERS ──────────────────────────────────────────────────────────────

def get_file_headers(file):
    """Read the first row headers and return them with auto-detect suggestions."""
    wb = _load_workbook(file)
    ws = wb.active
    first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not first_row:
        return {'headers': [], 'error': 'File appears to be empty.'}
    headers = [str(cell).strip() if cell is not None else f'Column {i+1}' for i, cell in enumerate(first_row)]
    # Auto-detect suggestions
    auto_map, _ = _detect_columns(ws)
    suggestions = {}
    for field, idx in auto_map.items():
        if idx is not None and idx < len(headers):
            suggestions[field] = headers[idx]
    return {'headers': headers, 'suggestions': suggestions}

# ─── IMPORT ────────────────────────────────────────────────────────────────────

def import_from_excel(file, mode='skip', column_mapping=None):
    """
    Import records from an Excel file.
    mode:
      'skip'   — skip records with duplicate (ticket, date) combinations
      'all'    — import everything regardless of duplicates
      'replace' — clear DB and re-import (destructive, SQLite only)
    column_mapping: optional dict like {'ticket': 'Header Name', ...} to override auto-detection
    """
    wb = _load_workbook(file)
    ws = wb.active

    if column_mapping:
        # Build col_map from user-provided mapping
        first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not first_row:
            raise Exception('Excel file appears to be empty.')
        header_to_idx = {}
        for i, cell in enumerate(first_row):
            if cell is not None:
                header_to_idx[str(cell).strip()] = i
        col_map = {}
        for field in ['ticket', 'query', 'executed_by', 'date', 'status', 'notes']:
            mapped_header = column_mapping.get(field, '')
            col_map[field] = header_to_idx.get(mapped_header)
        detected_headers = list(header_to_idx.keys())
    else:
        col_map, detected_headers = _detect_columns(ws)

    if col_map['ticket'] is None or col_map['query'] is None:
        raise Exception(
            f'Could not find required columns. '
            f'Detected columns: {detected_headers}. '
            f'Please ensure your file has "Service Ticket" and "Query Executed" columns.'
        )

    records, invalid_skipped = _extract_records(ws, col_map)

    if not records:
        return {
            'imported': 0, 'skipped': invalid_skipped, 'total': 0,
            'message': 'No valid records found in the Excel file.',
            'detected_columns': detected_headers
        }

    if mode == 'skip':
        existing = database.get_tickets_set()
        new_recs = [r for r in records if (r['ticket'], r['date']) not in existing]
        dup_skipped = len(records) - len(new_recs)
        imported = database.bulk_insert(new_recs) if new_recs else 0
        return {
            'imported': imported, 'skipped': dup_skipped + invalid_skipped,
            'total': len(records) + invalid_skipped,
            'message': f'Imported {imported} new records. Skipped {dup_skipped} duplicates.',
            'detected_columns': detected_headers, 'mode': 'skip'
        }
    else:
        imported = database.bulk_insert(records)
        return {
            'imported': imported, 'skipped': invalid_skipped,
            'total': len(records) + invalid_skipped,
            'message': f'Imported {imported} records.',
            'detected_columns': detected_headers, 'mode': 'all'
        }

# ─── SYNC PREVIEW ──────────────────────────────────────────────────────────────

def preview_sync(file):
    """Preview what an import would change without making any changes."""
    wb = _load_workbook(file)
    ws = wb.active
    try:
        col_map, detected_headers = _detect_columns(ws)
    except Exception as e:
        return {'error': str(e)}

    if col_map['ticket'] is None:
        return {
            'error': f'Could not detect ticket column. Found: {detected_headers}',
            'detected_columns': detected_headers
        }

    records, invalid = _extract_records(ws, col_map)
    existing = database.get_tickets_set()

    new_recs  = [r for r in records if (r['ticket'], r['date']) not in existing]
    dup_recs  = [r for r in records if (r['ticket'], r['date']) in existing]

    return {
        'total_in_file':   len(records) + invalid,
        'valid_rows':      len(records),
        'invalid_rows':    invalid,
        'new':             len(new_recs),
        'duplicates':      len(dup_recs),
        'will_import':     len(new_recs),
        'sample_new':      new_recs[:5],
        'detected_columns': detected_headers
    }

# ─── EXPORT: EXCEL ─────────────────────────────────────────────────────────────

def export_to_excel():
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise Exception('openpyxl not installed. Run: pip install openpyxl')

    records = database.get_all_for_export()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Backend Corrections'

    headers = ['ID', 'Service Ticket', 'Query Executed', 'Executed By', 'Date', 'Status', 'Notes', 'Created At']
    col_widths = [8, 22, 65, 20, 14, 20, 35, 22]

    # Header style
    hdr_fill = PatternFill(start_color='1e2640', end_color='1e2640', fill_type='solid')
    hdr_font = Font(bold=True, color='FFFFFF', name='Calibri', size=11)
    thin_side = Side(style='thin', color='2a3550')
    thin_border = Border(bottom=Side(style='medium', color='4f7ef8'))

    for col, (hdr, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col, value=hdr)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=False)
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.row_dimensions[1].height = 22

    status_fills = {
        'Completed':           PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid'),
        'Pending Verification':PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid'),
        'Rolled Back':         PatternFill(start_color='BDD7EE', end_color='BDD7EE', fill_type='solid'),
        'Failed':              PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid'),
    }

    for row_i, rec in enumerate(records, 2):
        values = [rec.get('id'), rec.get('ticket'), rec.get('query'),
                  rec.get('executed_by'), rec.get('date'), rec.get('status'),
                  rec.get('notes', ''), rec.get('created_at', '')]
        for col_i, val in enumerate(values, 1):
            cell = ws.cell(row=row_i, column=col_i, value=val)
            cell.alignment = Alignment(wrap_text=(col_i in [3, 7]), vertical='top')
            if row_i % 2 == 0:
                cell.fill = PatternFill(start_color='F8FAFF', end_color='F8FAFF', fill_type='solid')
        # Status colour
        status = rec.get('status', '')
        if status in status_fills:
            ws.cell(row=row_i, column=6).fill = status_fills[status]

    ws.freeze_panes = 'A2'
    if records:
        ws.auto_filter.ref = f'A1:H{len(records)+1}'

    os.makedirs(UPLOAD_DIR, exist_ok=True)
    filepath = os.path.join(UPLOAD_DIR, f'corrections_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx')
    wb.save(filepath)
    return filepath

# ─── EXPORT: CSV ───────────────────────────────────────────────────────────────

def export_to_csv():
    import csv, io as sio
    records = database.get_all_for_export()
    output = sio.StringIO()
    fieldnames = ['id', 'ticket', 'query', 'executed_by', 'date', 'status', 'notes', 'created_at']
    writer = csv.DictWriter(output, fieldnames=fieldnames, extrasaction='ignore')
    writer.writeheader()
    for rec in records:
        writer.writerow({k: rec.get(k, '') for k in fieldnames})
    return output.getvalue()

# ─── EXPORT: SHARED EXCEL ──────────────────────────────────────────────────────

def export_to_shared(filepath):
    """Export all corrections to a specific file path (shared drive / SharePoint sync folder).
    Overwrites the file each time to keep it in sync with the portal."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise Exception('openpyxl not installed.')

    records = database.get_all_for_export()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Backend Corrections'

    headers = ['ID', 'Service Ticket', 'Query Executed', 'Executed By', 'Date', 'Status', 'Notes', 'Created At']
    col_widths = [8, 22, 65, 20, 14, 20, 35, 22]

    hdr_fill = PatternFill(start_color='1e2640', end_color='1e2640', fill_type='solid')
    hdr_font = Font(bold=True, color='FFFFFF', name='Calibri', size=11)
    thin_border = Border(bottom=Side(style='medium', color='4f7ef8'))

    for col, (hdr, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col, value=hdr)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col)].width = width

    status_fills = {
        'Completed':           PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid'),
        'Pending Verification':PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid'),
        'Rolled Back':         PatternFill(start_color='BDD7EE', end_color='BDD7EE', fill_type='solid'),
        'Failed':              PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid'),
    }

    for row_i, rec in enumerate(records, 2):
        values = [rec.get('id'), rec.get('ticket'), rec.get('query'),
                  rec.get('executed_by'), rec.get('date'), rec.get('status'),
                  rec.get('notes', ''), rec.get('created_at', '')]
        for col_i, val in enumerate(values, 1):
            cell = ws.cell(row=row_i, column=col_i, value=val)
            cell.alignment = Alignment(wrap_text=(col_i in [3, 7]), vertical='top')
            if row_i % 2 == 0:
                cell.fill = PatternFill(start_color='F8FAFF', end_color='F8FAFF', fill_type='solid')
        status = rec.get('status', '')
        if status in status_fills:
            ws.cell(row=row_i, column=6).fill = status_fills[status]

    ws.freeze_panes = 'A2'
    if records:
        ws.auto_filter.ref = f'A1:H{len(records)+1}'

    # Write to temp file first, then replace (handles Office file locks)
    parent = os.path.dirname(filepath) or '.'
    os.makedirs(parent, exist_ok=True)

    import tempfile, time
    base = os.path.basename(filepath)
    tmp_fd, tmp_path = tempfile.mkstemp(suffix='.xlsx', prefix='~bcp_', dir=parent)
    os.close(tmp_fd)
    try:
        wb.save(tmp_path)
    except Exception as e:
        os.unlink(tmp_path)
        raise Exception(f'Failed to write temp file: {e}')

    # Try to replace the target file (retry if locked by Office / OneDrive)
    last_err = None
    for attempt in range(4):
        try:
            if os.path.exists(filepath):
                os.replace(tmp_path, filepath)
            else:
                os.rename(tmp_path, filepath)
            return filepath
        except PermissionError as e:
            last_err = e
            if attempt < 3:
                time.sleep(1)  # wait 1s and retry
    # All retries failed — save alongside with timestamp
    fallback = os.path.join(parent, f'corrections_sync_{datetime.now().strftime("%H%M%S")}.xlsx')
    try:
        os.rename(tmp_path, fallback)
    except Exception:
        pass
    raise Exception(
        f'Could not write to "{base}" — file may be open in Excel. '
        f'Data saved to "{os.path.basename(fallback)}" instead. '
        f'Close the file and try Sync again.'
    )
